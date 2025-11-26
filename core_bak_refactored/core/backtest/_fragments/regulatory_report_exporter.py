"""
监管合规报告格式导出工具（功能碎片）

业务目标5：监管合规报告能力
- 支持Excel、PDF、JSON API三种格式输出
- 字段完整性≥95%（20个必需字段）
- 符合巴塞尔III/沪深交易所指引格式要求

注意：
- 仅实现格式转换与导出逻辑，不引入业务判断
- PDF依赖reportlab（可选依赖，若无则跳过）
- Excel依赖openpyxl（开发环境已包含）
"""

import json
from typing import List, Optional
from pathlib import Path
import logging

from core_bak_refactored.core.backtest._fragments.stress_test_result import StressTestResult

logger = logging.getLogger(__name__)


class RegulatoryReportExporter:
    """
    监管报告导出器（多格式支持）
    
    支持格式：
    - Excel (.xlsx): 结构化表格，便于监管审查
    - JSON (.json): API接口标准格式
    - PDF (.pdf): 正式报告格式（可选，需reportlab）
    """
    
    @staticmethod
    def to_excel(results: List[StressTestResult], output_path: str) -> None:
        """
        导出为Excel格式（巴塞尔III/沪深交易所标准）
        
        Args:
            results: 压力测试结果列表
            output_path: 输出文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("openpyxl未安装，无法导出Excel格式")
            raise
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "压力测试报告"
        
        # 表头定义（20个必需字段）
        headers = [
            'report_id', 'portfolio_id', 'scenario_id',
            'var_normal', 'var_stressed',
            'stress_loss_amount', 'stress_loss_percentage',
            'recovery_period',
            'risk_decomposition',
            'triggered_actions', 'recommended_actions',
            'compliance_status',
            # 元数据字段
            'event_name', 'period',
            'predicted_loss', 'actual_loss', 'prediction_error',
            'benchmark_index'
        ]
        
        # 写入表头
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, result in enumerate(results, start=2):
            d = result.to_dict()
            meta = d.get('metadata', {}) or {}
            
            row_data = [
                d.get('report_id'),
                d.get('portfolio_id'),
                d.get('scenario_id'),
                d.get('var_normal'),
                d.get('var_stressed'),
                d.get('stress_loss_amount'),
                d.get('stress_loss_percentage'),
                d.get('recovery_period'),
                json.dumps(d.get('risk_decomposition', {}), ensure_ascii=False),
                json.dumps(d.get('triggered_actions', []), ensure_ascii=False),
                json.dumps(d.get('recommended_actions', []), ensure_ascii=False),
                d.get('compliance_status'),
                meta.get('event_name'),
                str(meta.get('period')) if meta.get('period') else None,
                meta.get('predicted_loss'),
                meta.get('actual_loss'),
                meta.get('prediction_error'),
                meta.get('benchmark_index'),
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 保存
        wb.save(output_path)
        logger.info(f"Excel报告已导出: {output_path}")
    
    @staticmethod
    def to_json(results: List[StressTestResult], output_path: str, indent: int = 2) -> None:
        """
        导出为JSON格式（API标准格式）
        
        Args:
            results: 压力测试结果列表
            output_path: 输出文件路径
            indent: JSON缩进空格数
        """
        data = {
            'report_type': 'stress_test',
            'version': '1.0',
            'total_scenarios': len(results),
            'results': [r.to_dict() for r in results]
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=indent)
        
        logger.info(f"JSON报告已导出: {output_path}")
    
    @staticmethod
    def to_pdf(results: List[StressTestResult], output_path: str) -> None:
        """
        导出为PDF格式（正式报告，可选功能）
        
        Args:
            results: 压力测试结果列表
            output_path: 输出文件路径
        
        Note:
            需要安装reportlab: pip install reportlab
            若未安装则跳过PDF生成
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            logger.warning("reportlab未安装，跳过PDF导出（可选功能）")
            return
        
        # 创建PDF文档（横向A4）
        doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=20,
        )
        
        # 标题
        title = Paragraph("压力测试监管合规报告", title_style)
        story.append(title)
        story.append(Spacer(1, 0.2*inch))
        
        # 摘要统计
        summary_style = styles['Normal']
        summary = Paragraph(f"场景数量: {len(results)}", summary_style)
        story.append(summary)
        story.append(Spacer(1, 0.3*inch))
        
        # 构建表格数据
        table_data = [
            ['Report ID', 'Portfolio', 'Scenario', 'Loss %', 'Error %', 'Compliance']
        ]
        
        for r in results:
            d = r.to_dict()
            table_data.append([
                d.get('report_id', '')[:8] + '...',  # 截断UUID
                d.get('portfolio_id', ''),
                d.get('scenario_id', ''),
                f"{d.get('stress_loss_percentage', 0):.2%}" if d.get('stress_loss_percentage') is not None else 'N/A',
                f"{d['metadata'].get('prediction_error', 0):.2%}" if d.get('metadata') and d['metadata'].get('prediction_error') is not None else 'N/A',
                d.get('compliance_status', 'N/A'),
            ])
        
        # 创建表格
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        
        # 生成PDF
        doc.build(story)
        logger.info(f"PDF报告已导出: {output_path}")
